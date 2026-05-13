from __future__ import annotations

from collections.abc import Iterable
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = Path("/Users/dmitriybocenkov/Desktop/Files/poker-bot-py.xlsx")
OUT_SQL_PATH = ROOT / "scripts" / "seed_from_excel.sql"


def _trim_headers(raw: list[object]) -> list[str]:
  last = 0
  for i, value in enumerate(raw, 1):
    if value is not None:
      last = i
  return [str(v) for v in raw[:last]]


def _sheet_rows(sheet) -> tuple[list[str], list[dict[str, object]]]:
  headers = _trim_headers([sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)])
  rows: list[dict[str, object]] = []
  for r in range(2, sheet.max_row + 1):
    values = [sheet.cell(r, c).value for c in range(1, len(headers) + 1)]
    if all(v is None for v in values):
      continue
    rows.append(dict(zip(headers, values, strict=True)))
  return headers, rows


def _to_int(value: object | None) -> int | None:
  if value is None or value == "":
    return None
  return int(value)


def _to_bool_int(value: object | None) -> int:
  return 1 if bool(int(value or 0)) else 0


def _to_date(value: object | None) -> str | None:
  if value is None or value == "":
    return None
  if isinstance(value, datetime):
    return value.date().isoformat()
  if isinstance(value, date):
    return value.isoformat()
  return str(value)


def _sql_value(value: object | None) -> str:
  if value is None:
    return "NULL"
  if isinstance(value, bool):
    return "1" if value else "0"
  if isinstance(value, (int, float)):
    if isinstance(value, float) and value.is_integer():
      return str(int(value))
    return str(value)
  escaped = str(value).replace("'", "''")
  return f"'{escaped}'"


def _insert_sql(table: str, columns: list[str], rows: Iterable[dict[str, object]]) -> list[str]:
  statements: list[str] = []
  for row in rows:
    values = ", ".join(_sql_value(row.get(c)) for c in columns)
    statements.append(f"INSERT INTO {table} ({', '.join(columns)}) VALUES ({values});")
  return statements


def main() -> None:
  wb = load_workbook(XLSX_PATH, data_only=True)
  data = {sheet_name: _sheet_rows(wb[sheet_name])[1] for sheet_name in wb.sheetnames}

  pokers_rows = data["pokers"]
  bet_tournament_param_rows: dict[str, dict[str, object]] = {}
  for r in data["bet_tournament_params"]:
    tournament_type = str(r.get("type") or "").strip()
    if not tournament_type:
      continue
    bet_tournament_param_rows[tournament_type] = {
      "row_id": _to_int(r.get("row_id")),
      "tournament_type": tournament_type,
      "bet_param_id": _to_int(r.get("bet_params_id")) or 1,
      "percent_to_first": _to_int(r.get("percent_to_first")) or 50,
      "percent_to_second": _to_int(r.get("percent_to_second")) or 30,
      "percent_to_third": _to_int(r.get("percent_to_third")) or 20,
      "duration_months": _to_int(r.get("duration_months")) or 12,
    }

  # Historical file has "regular/holiday". Current app expects "regular/year".
  if "holiday" in bet_tournament_param_rows and "year" not in bet_tournament_param_rows:
    holiday = dict(bet_tournament_param_rows["holiday"])
    holiday["tournament_type"] = "year"
    bet_tournament_param_rows["year"] = holiday

  # Keep only the tournament types supported by current app.
  bet_tournament_param_final = [r for key, r in bet_tournament_param_rows.items() if key in {"regular", "year"}]
  bet_tournament_param_final.sort(key=lambda x: x["row_id"] or 0)
  for i, row in enumerate(bet_tournament_param_final, 1):
    row["row_id"] = i

  bet_tournament_rows: list[dict[str, object]] = []
  # Seed minimal active banks expected by app.
  for idx, ttype in enumerate(("regular", "year"), 1):
    # Try to reuse bank values from historical rows by params_id relation.
    bank = 0
    if ttype == "regular":
      candidate = next((r for r in data["bet_tournaments"] if _to_int(r.get("params_id")) == 1), None)
    else:
      candidate = next((r for r in data["bet_tournaments"] if _to_int(r.get("params_id")) == 2), None)
    if candidate is not None:
      bank = _to_int(candidate.get("current_bank_size_kopecks")) or 0
    bet_tournament_rows.append(
      {
        "row_id": idx,
        "tournament_type": ttype,
        "current_bank_kopecks": bank,
        "params_id": 1,
      }
    )

  bet_rows: list[dict[str, object]] = []
  for r in data["bets"]:
    bet_date = _to_date(r.get("date"))
    if bet_date is None:
      continue
    bet_rows.append(
      {
        "row_id": _to_int(r.get("row_id")),
        "params_id": _to_int(r.get("params_id")) or 1,
        "date": bet_date,
        "better_name": r.get("better_name"),
        "better_id": _to_int(r.get("better_id")),
        "size_kopecks": _to_int(r.get("size_kopecks")) or 0,
        "winner": r.get("winner"),
        "looser": r.get("looser"),
        "score": _to_int(r.get("score")) or 0,
        "is_paid": _to_bool_int(r.get("is_paid")),
      }
    )

  poker_params_rows = []
  for r in data["poker_params"]:
    row_id = _to_int(r.get("row_id"))
    if row_id is None:
      continue
    poker_params_rows.append(
      {
        "row_id": row_id,
        "buyin_size_chips": _to_int(r.get("buyin_size_chips")) or 0,
        "buyin_size_kopecks": _to_int(r.get("buyin_size_kopecks")) or 0,
        "bb_size_chips": _to_int(r.get("bb_size_chips")) or 0,
        "max_buyins": _to_int(r.get("max_buyins")) or 0,
        "big_buyin": _to_int(r.get("big_buyin")),
        "big_buyin_pic": r.get("big_buyin_pic"),
        "super_buyin": _to_int(r.get("super_buyin")),
        "super_buyin_pic": r.get("super_buyin_pic"),
        "king_buyin": _to_int(r.get("king_buyin")),
        "king_buyin_pic": r.get("king_buyin_pic"),
      }
    )

  pokers_final = []
  for r in data["pokers"]:
    row_id = _to_int(r.get("row_id"))
    if row_id is None:
      continue
    pokers_final.append(
      {
        "row_id": row_id,
        "params_id": _to_int(r.get("params_id")) or 1,
        "date": _to_date(r.get("date")),
        "cashier_id": _to_int(r.get("cashier_id")),
        "is_going": _to_bool_int(r.get("is_going")),
        "is_bettable": _to_bool_int(r.get("is_bettable")),
        "is_ready_for_chips_entering": _to_bool_int(r.get("is_ready_for_chips_entering")),
        "winners": r.get("winners"),
        "loosers": r.get("loosers"),
      }
    )

  poker_data_final = []
  for r in data["poker_data"]:
    row_id = _to_int(r.get("row_id"))
    if row_id is None:
      continue
    poker_data_final.append(
      {
        "row_id": row_id,
        "date": _to_date(r.get("date")),
        "player_name": r.get("player_name"),
        "player_id": _to_int(r.get("player_id")),
        "is_prev_winner": _to_bool_int(r.get("is_prev_winner")),
        "buyins": _to_int(r.get("buyins")) or 0,
        "big_buyin_count": _to_int(r.get("big_buyin_count")) or 0,
        "super_buyin_count": _to_int(r.get("super_buyin_count")) or 0,
        "chips": _to_int(r.get("chips")) or 0,
        "money_kopecks": _to_int(r.get("money_kopecks")) or 0,
      }
    )

  buyins_data_final = []
  for r in data["buyins_data"]:
    row_id = _to_int(r.get("row_id"))
    if row_id is None:
      continue
    buyins_data_final.append(
      {
        "row_id": row_id,
        "date": _to_date(r.get("date")),
        "player_id": _to_int(r.get("player_id")),
        "player_name": r.get("player_name"),
        "buyin": _to_int(r.get("buyin")) or 0,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
      }
    )

  bet_params_final = []
  for r in data["bet_params"]:
    row_id = _to_int(r.get("row_id"))
    if row_id is None:
      continue
    if r.get("small_size_kopecks") is None:
      continue
    bet_params_final.append(
      {
        "row_id": row_id,
        "small_pic": r.get("small_pic") or "🐤",
        "small_size_kopecks": _to_int(r.get("small_size_kopecks")) or 0,
        "small_score": _to_int(r.get("small_score")) or 0,
        "small_score_combo": _to_int(r.get("small_score_combo")) or 0,
        "big_pic": r.get("big_pic") or "🐔",
        "big_size_kopecks": _to_int(r.get("big_size_kopecks")) or 0,
        "big_score": _to_int(r.get("big_score")) or 0,
        "big_score_combo": _to_int(r.get("big_score_combo")) or 0,
        "percent_to_regular_bank_if_it_is_going": _to_int(r.get("percent_to_regular_bank_if_it_is_going")) or 0,
      }
    )

  stat_indicators_final = []
  for r in data["stat_indicators"]:
    row_id = _to_int(r.get("row_id"))
    if row_id is None:
      continue
    stat_indicators_final.append(
      {
        "row_id": row_id,
        "type": r.get("type"),
        "description": r.get("description"),
        "description_full": r.get("description_full"),
        "pic": r.get("pic"),
        "for_current_tournaments": r.get("for_current_tournaments"),
        "is_for_achievement": _to_bool_int(r.get("is_for_achievement")),
      }
    )

  achievements_final = []
  for r in data["achievements"]:
    row_id = _to_int(r.get("row_id"))
    if row_id is None:
      continue
    achievements_final.append(
      {
        "row_id": row_id,
        "type": r.get("type"),
        "sort": r.get("sort"),
        "description": r.get("description"),
        "pic": r.get("pic"),
        "stat_id": _to_int(r.get("stat_id")) or 0,
        "is_permanent": _to_bool_int(r.get("is_permanent")),
      }
    )

  sql_lines: list[str] = [
    "BEGIN;",
    "PRAGMA foreign_keys = OFF;",
    "",
    "-- Generated from poker-bot-py.xlsx (users excluded)",
  ]

  # Delete in dependency-safe order.
  sql_lines.extend(
    [
      "DELETE FROM achievements;",
      "DELETE FROM stat_indicators;",
      "DELETE FROM buyins_data;",
      "DELETE FROM bets;",
      "DELETE FROM poker_data;",
      "DELETE FROM pokers;",
      "DELETE FROM poker_params;",
      "DELETE FROM bet_tournaments;",
      "DELETE FROM bet_tournament_params;",
      "DELETE FROM bet_params;",
      "DELETE FROM poker_room_denied;",
      "",
    ]
  )

  sql_lines += _insert_sql(
    "poker_params",
    [
      "row_id",
      "buyin_size_chips",
      "buyin_size_kopecks",
      "bb_size_chips",
      "max_buyins",
      "big_buyin",
      "big_buyin_pic",
      "super_buyin",
      "super_buyin_pic",
      "king_buyin",
      "king_buyin_pic",
    ],
    poker_params_rows,
  )
  sql_lines.append("")

  sql_lines += _insert_sql(
    "pokers",
    [
      "row_id",
      "params_id",
      "date",
      "cashier_id",
      "is_going",
      "is_bettable",
      "is_ready_for_chips_entering",
      "winners",
      "loosers",
    ],
    pokers_final,
  )
  sql_lines.append("")

  sql_lines += _insert_sql(
    "poker_data",
    [
      "row_id",
      "date",
      "player_name",
      "player_id",
      "is_prev_winner",
      "buyins",
      "big_buyin_count",
      "super_buyin_count",
      "chips",
      "money_kopecks",
    ],
    poker_data_final,
  )
  sql_lines.append("")

  sql_lines += _insert_sql(
    "buyins_data",
    ["row_id", "date", "player_id", "player_name", "buyin", "created_at"],
    buyins_data_final,
  )
  sql_lines.append("")

  sql_lines += _insert_sql(
    "bet_params",
    [
      "row_id",
      "small_pic",
      "small_size_kopecks",
      "small_score",
      "small_score_combo",
      "big_pic",
      "big_size_kopecks",
      "big_score",
      "big_score_combo",
      "percent_to_regular_bank_if_it_is_going",
    ],
    bet_params_final,
  )
  sql_lines.append("")

  sql_lines += _insert_sql(
    "bet_tournament_params",
    [
      "row_id",
      "tournament_type",
      "bet_param_id",
      "percent_to_first",
      "percent_to_second",
      "percent_to_third",
      "duration_months",
    ],
    bet_tournament_param_final,
  )
  sql_lines.append("")

  sql_lines += _insert_sql(
    "bet_tournaments",
    ["row_id", "tournament_type", "current_bank_kopecks", "params_id"],
    bet_tournament_rows,
  )
  sql_lines.append("")

  sql_lines += _insert_sql(
    "bets",
    [
      "row_id",
      "params_id",
      "date",
      "better_name",
      "better_id",
      "size_kopecks",
      "winner",
      "looser",
      "score",
      "is_paid",
    ],
    bet_rows,
  )
  sql_lines.append("")

  sql_lines += _insert_sql(
    "stat_indicators",
    ["row_id", "type", "description", "description_full", "pic", "for_current_tournaments", "is_for_achievement"],
    stat_indicators_final,
  )
  sql_lines.append("")

  sql_lines += _insert_sql(
    "achievements",
    ["row_id", "type", "sort", "description", "pic", "stat_id", "is_permanent"],
    achievements_final,
  )
  sql_lines.extend(["", "PRAGMA foreign_keys = ON;", "COMMIT;"])

  OUT_SQL_PATH.write_text("\n".join(sql_lines) + "\n", encoding="utf-8")
  print(f"Generated: {OUT_SQL_PATH}")
  print(f"Rows: pokers={len(pokers_final)} poker_data={len(poker_data_final)} bets={len(bet_rows)}")


if __name__ == "__main__":
  main()
